#!/usr/bin/env python
# coding: utf-8

# In[1]:


# LIBRERIAS -------------------------------------------------------------------------------------
import tkinter as tk
from tkinterdnd2 import TkinterDnD, DND_FILES
import openpyxl
import os
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from datetime import datetime
import io


# FUNCIÓN PARA ARRASTRAR ARCHIVOS A LA VENTANA ----------------------------------------------------
#Se me pidió en la empresa que pudiera hacer una aplicación simple en la que pudieran solo arrastrar el archivo y que automáticamente se hagan todos los cambios

def dropear(event):
    ruta = event.data.strip('{}')
    entry.delete(0, tk.END)
    entry.insert(0, ruta)
    print(f"La ruta de archivo es: {ruta}")

# FUNCIÓN PARA HACER LA TABLA TRANSPUESTA INPUT =====================================================================
# La idea de esta funcion es tomar los valores de BFSDATA y transponerlos y transformarlos al formato que se ocupa para alimentar cierto archivo de Power BI que maneja el equipo de Demmand
# Por temas de confidencialidad, no puedo mostrar el BI :C (no me dieron permiso)

def hojaresumen(ruta):
    wb = openpyxl.load_workbook(ruta)
    ws = wb['BFSDATA']
    
    #Algunas variables que se ocuparan en la tabla
    nombre_archivo = os.path.basename(ruta)
    nombre_archivo = os.path.splitext(nombre_archivo)[0]
    nombre_archivo_split = nombre_archivo.split(" ")
    
    distribuidor = nombre_archivo_split[0]
    fecha_sn = nombre_archivo_split[2]
    fecha_sn = "01-" + str(fecha_sn)
    
    encabezados = [cell.value for cell in ws[1]]
    
    #Elementos de BFSDATA a ocupar en la tabla
    cod_sap_col_idx = encabezados.index("COD SAP")
    canal_comercial_col_idx = encabezados.index("CANAL COMERCIAL")
    cliente_col_idx = encabezados.index("CLIENTE")
    
    columnas_adicionales = encabezados[9:]
    
    datos = []
    
    # Generalmente, ellos ocupan dos columnas, una de Canal Comercial y otra solo de Cliente
    # La regla que ellos ocupan a la hora de hacer reportes es que CC toma los valores de Canal_Comercial, pero cuando es "Supermercado", toman los valores de "Cliente"
    #Esto es por una cuestion de jerarquias de como ellos han estructurado los datos
    
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_sap = row[cod_sap_col_idx]
        if row[canal_comercial_col_idx] == "SUPERMERCADO":
            canal_comercial = row[cliente_col_idx]
        else:
            canal_comercial = row[canal_comercial_col_idx]
        
        valores_adicionales = list(row)[9:]
        
        # Formatear FECHAS> Las fechas vienen dadas por un formato "ene-2024", ellos me pidieron que la transformacion lo cambie a un formato "01-01-2024"
        #El día es irrelevante, por eso siempre es 01
        
        meses = {
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
        }
            
        valores_fmt = []
        for val in columnas_adicionales:
            partes = val.split('-')
            año = int(partes[0])
            mes_abreviado = partes[1].lower()
            mes_numero = meses[mes_abreviado]
            fecha = datetime(año, mes_numero, 1)
            fecha_formateada = fecha.strftime("%d-%m-%Y")
            valores_fmt.append(fecha_formateada)
        
        for idx, valor in enumerate(valores_adicionales):
            datos.append([cod_sap, distribuidor, canal_comercial, fecha_sn, valores_fmt[idx], valor])
    
    #Aarmar el dataframe que se ocupara en la hoja INPUT
    
    df = pd.DataFrame(datos, columns=["COD SAP", "DISTRIBUIDOR", "CANAL COMERCIAL", "FECHA SN", "MES", "FCST"])
    df['COD'] = df['COD SAP'].str[:9]
    df['DESCRIPCIÓN'] = df['COD SAP'].str[10:]
    df = df[["COD", "DESCRIPCIÓN", "DISTRIBUIDOR", "CANAL COMERCIAL", "FECHA SN", "MES", "FCST"]]
    
    return df, wb

#FUNCIÓN PARA DAR FORMATO AL ARCHIVO PARA EL EQUIPO DE BI ===================================================================
#La segunda tarea que me pidieron para este programa es hacer que el archivo que arrastres, el programa aplique una transformación
#De tal manera que haga una tabla pivote para cada uno de los clientes/canales_comerciales

#Es decir, por cada Cliente y Canal Comercial, hay una hoja, con la respectiva suma de los productos, para todas las fechas del periodo 

#Esto se hace de esta manera, porque es el formato que el equipo de Demand Planning da al equipo de Business Intelligence para cargar la información

def darformatoBI(wb, ruta):
    ws = wb['BFSDATA']
    for row in ws.iter_rows():
        for cell in row: #Correccion de algunos errores comunes en SN (errores de dedo o errores que se dan al generar el SN)
            if cell.value == "ERR":
                cell.value = 0
            elif cell.value == " Pronóstico":
                cell.value = "Pronósticos"
                
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)

    expected_columns = ["COD SAP", "Tipo de registro", "Unidades"]
    for col in expected_columns:
        if col not in df.columns:
            raise KeyError(f"La columna esperada '{col}' no se encuentra en los datos del archivo Excel") #Un precautionary error 

   #Mismo procedimiento que en la función anterior

    valores = df.columns[9:]
    meses = {
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }

    valores_fmt = []

    for val in valores:
        partes = val.split('-')
        año = int(partes[0])
        mes_abreviado = partes[1].lower()
        mes_numero = meses[mes_abreviado]
        fecha = datetime(año, mes_numero, 1)
        fecha_formateada = fecha.strftime("%Y/%m/%d")
        valores_fmt.append(fecha_formateada)

    nuevos_nombres = {df.columns[i + 9]: val for i, val in enumerate(valores_fmt)}
    df.rename(columns=nuevos_nombres, inplace=True)
    
    valores = df.columns[9:]
    lista_cc = df['CANAL COMERCIAL'].drop_duplicates()
    lista_cliente = df['CLIENTE'].drop_duplicates()
    
    df['COD'] = df['COD SAP'].str[:9]
    df['DESCRIPCIÓN'] = df['COD SAP'].str[10:]
    
    pivots_cliente = {}
    pivots_cc = {}
    
    #Retorna los pivotes por cliente y cc 
    
    for cc in lista_cc:
        for cliente in lista_cliente:
            df1 = df[(df['CANAL COMERCIAL'] == cc) & (df['CLIENTE'] == cliente)]
            if not df1.empty:
                pivot = df1.pivot_table(index=['COD', 'DESCRIPCIÓN'], values=valores, aggfunc='sum').reset_index()
                pivots_cliente[cliente] = pivot
                pivots_cc[cc] = pivot
    
    return pivots_cliente, pivots_cc

#Esta funcion es para llamar las funciones anteriores, pero ya ocupando las herramientas de os
def call_funciones():
    ruta = entry.get()
    if os.path.isfile(ruta):
        global df, original_ruta, original_wb, pivots_cliente, pivots_cc
        df, original_wb = hojaresumen(ruta)
        pivots_cliente, pivots_cc = darformatoBI(original_wb, ruta)
        original_ruta = ruta
        mostrar_df(df)
        plot_canal_comercial(df)
    else:
        print("Ingrese otra ruta de archivo, la actual es invalida")

#Para que en la ventana emergente muestre el df
def mostrar_df(df):
    text_widget.delete(1.0, tk.END)
    text_widget.insert(tk.END, df.to_string(index=False))

#FUNCIÓN PARA GUARDAR LOS CAMBIOS EN UNA NUEVA HOJA DE EXCEL =================================================================
def guardar_excel():
    if df is None:
        messagebox.showerror("Error", "No hay datos para guardar. Procesa un archivo primero.")
        return
    
    #base_ruta es para que por default sugiera un nombre del archivo nuevo al guardarlo, para que sea dentro de la misma ruta del original
    base_ruta = os.path.splitext(original_ruta)[0] + "_1.xlsx"
    ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=base_ruta, filetypes=[("Excel files", "*.xlsx")])
    if ruta:
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df.to_excel(writer, nombre_hoja="INPUT", index=False)
        
        #Volver a copiar BFSDATA en el nuevo archivo
        new_wb = openpyxl.load_workbook(ruta)
        new_ws = new_wb.create_sheet(title='BFSDATA')
        
        original_ws = original_wb['BFSDATA']
        for row in original_ws.iter_rows():
            for cell in row:
                new_ws[cell.coordinate].value = cell.value
        
        # Guardar las tablas pivote 
        for cliente, pivot in pivots_cliente.items():
            nombre_hoja = f"{cliente[:30]}"  # No van a haver más de 30 categorías 
            new_ws = new_wb.create_sheet(title=nombre_hoja)
            
            #Encabezados 
            encab = pivot.columns.tolist()
            for col_idx, header in enumerate(encab, 1):
                new_ws.cell(row=1, column=col_idx, value=header)
            # Datos
            for row_idx, row in enumerate(pivot.itertuples(index=False), 2): #Encontré información de las intertuplas en: https://www.programiz.com/python-programming/pandas/methods/itertuples
                for col_idx, value in enumerate(row, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)
        
        for cc, pivot in pivots_cc.items():
            nombre_hoja = f"{cc[:30]}"  
            print(f"Saving pivot table for channel: {cc} as {nombre_hoja}")
            new_ws = new_wb.create_sheet(title=nombre_hoja)
            # Encabezados
            encab = pivot.columns.tolist()
            for col_idx, header in enumerate(encab, 1):
                new_ws.cell(row=1, column=col_idx, value=header)
            # Datos
            for row_idx, row in enumerate(pivot.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # GUARDAR ---------------------------------------------------------------------------------------------
        new_wb.save(ruta)
        messagebox.showinfo("Guardado", f"Archivo guardado exitosamente en {ruta}")

#FUNCIÓN PARA CREAR EL GRÁFICO ===============================================================================================
#Finalmente, el equipo me pidió que la herramienta creara un gráfico para ver rápidamente cuál de los clientes/canales_comerciales tendría mayor FCST dentro del periodo


def plot_canal_comercial(df):
    # AGRUPAR Y SUMAR
    df_grouped = df.groupby('CANAL COMERCIAL')['FCST'].sum().reset_index()
    
    # Crear gráfico de barras
    plt.figure(figsize=(9, 6))
    plt.bar(df_grouped['CANAL COMERCIAL'], df_grouped['FCST'])
    plt.xlabel('Canal Comercial')
    plt.ylabel('Suma de FCST')
    plt.title('Suma de FCST por Canal Comercial')
    plt.xticks(rotation=45, ha='right')

    # Almacenar el gráfico y para luego usarlo dentro del Tk
    img_stream = io.BytesIO()
    plt.savefig(img_stream, format='png')
    img_stream.seek(0)
    plt.close()

    img = Image.open(img_stream)
    img_tk = ImageTk.PhotoImage(img)

    # Ajustar tamaño lbl 
    img_label.config(image=img_tk, width=img_tk.width(), height=img_tk.height())
    img_label.image = img_tk

#VENTANA DE LA APLICACIÓN ---------------------
root = TkinterDnD.Tk()
root.title("McCormick - Programa para formato SN")
root.geometry("900x800")

# Frame para el área de entrada y botones
frame_top = tk.Frame(root)
frame_top.pack(fill=tk.X, pady=10)

# Caja donde se ve el nombre del archivo que arrastras - o Entry para hacerlo más elegante
entry = tk.Entry(frame_top, width=40)
entry.pack(side=tk.LEFT, padx=10)

#Botón para llamar las funciones
btn = tk.Button(frame_top, text="Procesar Archivo", command=call_funciones)
btn.pack(side=tk.LEFT, padx=10)

#Botón para guardar los cambios en un nuevo excel
guardar_btn = tk.Button(frame_top, text="Guardar como Excel", command=guardar_excel)
guardar_btn.pack(side=tk.LEFT, padx=10)

#Frame para el área de texto (Esta es la tabla para visualizar el df transpuesto de la función "hojaresumen")--------
frame_text = tk.Frame(root)
frame_text.pack(fill=tk.BOTH, expand=True, pady=10)
text_widget = tk.Text(frame_text, wrap='none', width=100, height=5)
text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

#Frame para el área del grafico (Donde se ve la img del gráfico) -----------------------------------------------------
frame_img = tk.Frame(root)
frame_img.pack(fill=tk.BOTH, expand=True, pady=0)

img_label = tk.Label(frame_img)
img_label.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=0, pady=0)

# ---------dropear arch----

root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', dropear)

root.mainloop()

